"""读《微调数据集设计总表》，并校验它自身的一致性。

这张表是数据集的唯一事实来源：Tool → Parameter → 能力/规律 → Scenario → State → Variable。
数据只从表生成，不手改；表改了就重跑流水线。

表里的受控词表从 00_说明与字典 的「字典区」动态读取，不在代码里硬编码 ——
否则表改了代码不知道，就又出现两个事实来源。

用法：
    PYTHONPATH=dataset python3 -m esa.design_table --check
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

from .ir import load_schemas, schemas_by_name
from .paths import in_dataset

DEFAULT_XLSX = Path.home() / "Downloads" / "微调数据集设计总表.xlsx"
DEFAULT_SCHEMAS = str(in_dataset("schemas/tool_schemas.json"))

# 每张表的首列名，用来定位真正的表头行（各表标题/提示行数不一致）
SHEET_KEYS = {
    "01_Tool": "tool_id",
    "02_Parameter": "parameter_id",
    "03_能力与规律": "ability_id",
    "04_Scenario": "scenario_id",
    "05_State": "state_id",
    "06_Variable": "variable_value_id",
    "07_审核": "review_id",
}

# 00 页字典区的表头行
DICT_HEADER_ROW = 14


@dataclass
class Issue:
    level: str  # error | warn
    where: str
    detail: str

    def __str__(self) -> str:
        mark = "❌" if self.level == "error" else "⚠️ "
        return f"{mark} [{self.where}] {self.detail}"


def _find_header_row(ws, key: str) -> int:
    for ri in range(1, min(ws.max_row, 12) + 1):
        for cell in ws[ri]:
            if cell.value is not None and str(cell.value).strip() == key:
                return ri
    raise ValueError(f"{ws.title}: 找不到表头行（期望首列 {key!r}）")


def _read_sheet(ws, key: str) -> list[dict[str, Any]]:
    hr = _find_header_row(ws, key)
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[hr]]
    rows = []
    for raw in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not raw or raw[0] is None or not str(raw[0]).strip():
            continue
        rec = {}
        for i, name in enumerate(header):
            if not name:
                continue
            v = raw[i] if i < len(raw) else None
            rec[name] = str(v).strip() if isinstance(v, str) else v
        rows.append(rec)
    return rows


def _read_dictionary(ws) -> dict[str, list[str]]:
    """字典区是按列排的受控词表：第 14 行是字段名，往下是候选值。"""
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[DICT_HEADER_ROW]]
    vocab: dict[str, list[str]] = {}
    for ci, name in enumerate(header):
        if not name or "\n" in name:  # 最后一列塞的是决策树说明，不是词表
            continue
        vals = []
        for ri in range(DICT_HEADER_ROW + 1, ws.max_row + 1):
            v = ws.cell(row=ri, column=ci + 1).value
            if v is not None and str(v).strip():
                vals.append(str(v).strip())
        if vals:
            vocab[name] = vals
    return vocab


@dataclass
class DesignTable:
    tools: list[dict] = field(default_factory=list)
    parameters: list[dict] = field(default_factory=list)
    rules: list[dict] = field(default_factory=list)
    scenarios: list[dict] = field(default_factory=list)
    states: list[dict] = field(default_factory=list)
    variables: list[dict] = field(default_factory=list)
    reviews: list[dict] = field(default_factory=list)
    vocab: dict[str, list[str]] = field(default_factory=dict)

    @classmethod
    def load(cls, path: str | Path = DEFAULT_XLSX) -> DesignTable:
        wb = openpyxl.load_workbook(path, data_only=True)
        get = lambda name: _read_sheet(wb[name], SHEET_KEYS[name])  # noqa: E731
        return cls(
            tools=get("01_Tool"),
            parameters=get("02_Parameter"),
            rules=get("03_能力与规律"),
            scenarios=get("04_Scenario"),
            states=get("05_State"),
            variables=get("06_Variable"),
            reviews=get("07_审核"),
            vocab=_read_dictionary(wb["00_说明与字典"]),
        )

    def variables_by_state(self) -> dict[str, dict[str, list]]:
        out: dict[str, dict[str, list]] = {}
        for v in self.variables:
            if str(v.get("enabled", "是")) != "是":
                continue
            out.setdefault(str(v["state_id"]), {}).setdefault(str(v["variable_name"]), []).append(v["value"])
        return out

    def combo_count(self, state_id: str) -> int:
        """该 State 的事实/表达变量笛卡尔积规模。没有变量池时返回 0。"""
        vs = self.variables_by_state().get(state_id)
        if not vs:
            return 0
        n = 1
        for vals in vs.values():
            n *= len(vals)
        return n


# --------------------------------------------------------------------------
# 一致性校验
# --------------------------------------------------------------------------


def check(dt: DesignTable, schemas_path: str | Path = DEFAULT_SCHEMAS) -> list[Issue]:
    out: list[Issue] = []
    actions = set(dt.vocab.get("expected_action", []))

    rule_ids = {str(r["rule_id"]) for r in dt.rules if r.get("rule_id")}
    scen_ids = {str(s["scenario_id"]) for s in dt.scenarios if s.get("scenario_id")}
    state_ids = {str(s["state_id"]) for s in dt.states if s.get("state_id")}

    # --- 受控词表 ---
    for r in dt.rules:
        a = r.get("expected_action")
        if a and str(a) not in actions:
            out.append(
                Issue("error", f"03_能力与规律/{r['rule_id']}", f"expected_action={a!r} 不在受控词表 {sorted(actions)}")
            )
    for s in dt.states:
        a = s.get("expected_action")
        if a and str(a) not in actions:
            out.append(Issue("error", f"05_State/{s['state_id']}", f"expected_action={a!r} 不在受控词表"))

    # --- 规则与状态的标签是否打架 ---
    # 规则的 expected_action 描述的是「触发条件满足且 gate 通过」时的动作。
    # 负样本（gate 未满足 / Hard Negative）本来就该给出不同动作，不算冲突。
    by_rule = {str(r["rule_id"]): r for r in dt.rules if r.get("rule_id")}
    for s in dt.states:
        rid, sa = str(s.get("primary_rule_id") or ""), s.get("expected_action")
        ra = by_rule.get(rid, {}).get("expected_action")
        if not (rid and sa and ra) or str(sa) == str(ra):
            continue
        if str(s.get("case_type") or "") in ("Hard Negative", "异常案例") or str(s.get("gate_state") or "") == "未满足":
            continue
        out.append(
            Issue("error", f"05_State/{s['state_id']}", f"动作={sa}，但所属规则 {rid} 声明的是 {ra} —— 生成器不知道听谁的")
        )

    # --- 外键 ---
    for s in dt.states:
        if (rid := str(s.get("primary_rule_id") or "")) and rid not in rule_ids:
            out.append(Issue("error", f"05_State/{s['state_id']}", f"引用了不存在的规则 {rid}"))
        if (sid := str(s.get("scenario_id") or "")) and sid not in scen_ids:
            out.append(Issue("error", f"05_State/{s['state_id']}", f"引用了不存在的场景 {sid}"))
    for v in dt.variables:
        if (sid := str(v.get("state_id") or "")) and sid not in state_ids:
            out.append(Issue("error", f"06_Variable/{v['variable_value_id']}", f"引用了不存在的状态 {sid}"))

    # --- 每条规律至少要有一个 State，否则等于没落地 ---
    covered = {str(s.get("primary_rule_id")) for s in dt.states if s.get("primary_rule_id")}
    for rid in sorted(rule_ids - covered):
        name = by_rule[rid].get("rule_name", "")
        out.append(Issue("error", f"03_能力与规律/{rid}", f"「{name}」有定义但零个 State，规律没有落地"))

    # --- State 完整性 ---
    for s in dt.states:
        sid = s["state_id"]
        if not s.get("expected_action"):
            out.append(Issue("error", f"05_State/{sid}", "缺 expected_action —— 无法生成 gold label"))
        if not s.get("state_name"):
            out.append(Issue("error", f"05_State/{sid}", "缺 state_name"))
        # gate 两列必须成对
        gc, gs = s.get("gate_condition"), s.get("gate_state")
        if bool(gc) != bool(gs):
            out.append(Issue("warn", f"05_State/{sid}", "gate_condition 与 gate_state 只填了一个"))
        # 缺参数状态与 missing_parameters 必须自洽
        rps, miss = str(s.get("required_param_state") or ""), str(s.get("missing_parameters") or "")
        if rps in ("缺1个", "缺多个") and not miss:
            out.append(Issue("error", f"05_State/{sid}", f"required_param_state={rps} 但 missing_parameters 为空"))
        if rps == "全部齐全" and miss:
            out.append(Issue("error", f"05_State/{sid}", f"required_param_state=全部齐全 但列了缺失参数 {miss!r}"))

    # --- Variable 的核心不变量 ---
    for v in dt.variables:
        if str(v.get("affects_gold_answer", "否")) != "否":
            out.append(
                Issue("error", f"06_Variable/{v['variable_value_id']}", "affects_gold_answer 必须为「否」，会改变答案的应提升为 State")
            )

    # --- 与 tool_schemas.json 对齐 ---
    try:
        schemas, version = load_schemas(schemas_path)
        by_name = schemas_by_name(schemas)
        table_tools = {str(t["tool_name"]) for t in dt.tools if t.get("tool_name")}
        for extra in sorted(table_tools - set(by_name)):
            out.append(Issue("error", "01_Tool", f"表里有 {extra!r}，但 tool_schemas.json 里没有"))
        for missing in sorted(set(by_name) - table_tools):
            out.append(Issue("warn", "01_Tool", f"tool_schemas.json 里有 {missing!r}，但表里没登记"))
        for s in dt.states:
            et = str(s.get("expected_tool") or "")
            for name in re.split(r"[/,、\s]+", et):
                if name and name not in by_name:
                    out.append(Issue("error", f"05_State/{s['state_id']}", f"expected_tool={name!r} 不存在"))
    except FileNotFoundError:
        out.append(Issue("warn", "schemas", f"找不到 {schemas_path}，跳过 schema 对齐检查"))
        version = "?"

    # --- 预算 vs 变量多样性 ---
    for s in dt.states:
        sid = str(s["state_id"])
        budget = s.get("generation_budget") or 0
        combos = dt.combo_count(sid)
        if budget and combos == 0:
            out.append(Issue("warn", f"05_State/{sid}", f"预算 {budget} 条，但没有任何 Variable 候选值"))
        elif budget and combos and budget / combos > 3:
            out.append(
                Issue("warn", f"05_State/{sid}", f"预算 {budget} / 组合 {combos} = 重复 {budget / combos:.1f}x，超过 3x 上限")
            )

    return out


def summarize(dt: DesignTable) -> None:
    from collections import Counter

    print(f"Tool {len(dt.tools)} | Parameter {len(dt.parameters)} | 规律 {len(dt.rules)} | "
          f"Scenario {len(dt.scenarios)} | State {len(dt.states)} | Variable {len(dt.variables)}")
    acts = Counter(str(s.get("expected_action") or "(空)") for s in dt.states)
    budget = Counter()
    for s in dt.states:
        budget[str(s.get("expected_action") or "(空)")] += s.get("generation_budget") or 0
    total = sum(budget.values()) or 1
    print(f"\n{'expected_action':22s} {'State数':>6s} {'预算':>6s} {'占比':>7s}")
    for a, n in acts.most_common():
        print(f"{a:22s} {n:6d} {budget[a]:6d} {budget[a] / total * 100:6.1f}%")
    print(f"{'合计':22s} {len(dt.states):6d} {total:6d}")


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="读取并校验设计总表")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--schemas", default=DEFAULT_SCHEMAS)
    ap.add_argument("--check", action="store_true", help="校验失败时以非零退出码结束")
    args = ap.parse_args(argv)

    dt = DesignTable.load(args.xlsx)
    summarize(dt)

    issues = check(dt, args.schemas)
    errors = [i for i in issues if i.level == "error"]
    warns = [i for i in issues if i.level == "warn"]

    print(f"\n{'=' * 70}\n{len(errors)} 个错误 / {len(warns)} 个警告\n{'=' * 70}")
    for i in errors + warns:
        print(" ", i)

    return 1 if (args.check and errors) else 0


if __name__ == "__main__":
    sys.exit(main())
